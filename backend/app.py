from flask import Flask, request, jsonify, send_from_directory
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging

app = Flask(__name__, static_folder='build', static_url_path='')

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Define the path for the Excel file within the data directory
excel_file_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'data', 'emails.xlsx')

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

@app.route('/submit_email', methods=['POST'])
def submit_email():
    email = request.form['email']
    logging.debug(f"Received email: {email}")

    try:
        # Check if the Excel file exists; if not, create it with headers
        if not os.path.exists(excel_file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Emails"
            # Set headers
            headers = ["Index", "Email", "Date & Time"]
            sheet.append(headers)
            # Apply border to headers
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            workbook.save(excel_file_path)
            logging.debug("Created Excel file and wrote headers")

        # Load the workbook and select the active sheet
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # Find the first available row
        next_index = None
        for row in range(2, sheet.max_row + 2):
            if sheet.cell(row=row, column=1).value is None:
                next_index = row
                break

        if next_index is None:
            next_index = sheet.max_row + 1

        # Get the current date and time
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Prepare the data
        row_data = [next_index - 1, email, current_time]

        # Write the data to the first available row
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=next_index, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(excel_file_path)
        logging.debug("Appended email to Excel file with index and date-time")
    except Exception as e:
        logging.error(f"Error writing to Excel file: {e}", exc_info=True)
        return jsonify({'message': 'Failed to submit email'}), 500

    return jsonify({'message': 'Email submitted successfully'})

@app.route('/send_feedback', methods=['POST'])
def send_feedback():
    data = request.get_json()
    feedback_message = data.get('message')

    try:
        # Set up the server
        server = smtplib.SMTP('smtp.gmail.com', 587) # Change this to your email provider's SMTP server
        server.starttls()

        # Login to your email account
        server.login('medgenieai@gmail.com', 'mqtq wfjd gory yrza') # Change to your email and password

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = 'alexfarouz@gmail.com'  # Feedback email address
        msg['To'] = 'medgenieai@gmail.com'  # Your email
        msg['Subject'] = 'Feedback from MedGenie.ai'
        
        body = f"Feedback:\n\n{feedback_message}"
        msg.attach(MIMEText(body, 'plain'))

        # Send the email
        server.send_message(msg)
        server.quit()

        return jsonify({'message': 'Feedback sent successfully!'}), 200
    except Exception as e:
        logging.error(f"Error sending feedback email: {e}", exc_info=True)
        return jsonify({'message': 'Failed to send feedback.'}), 500

@app.route('/')
def serve():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def static_proxy(path):
    # Send any other paths to the static folder
    return send_from_directory(app.static_folder, path)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
